import json
import os
import psycopg2
from typing import Dict, Any
from io import BytesIO
import base64

def handler(event: Dict[str, Any], context: Any) -> Dict[str, Any]:
    '''
    Business: Экспорт каналов в Excel (XLSX) и CSV форматы
    Args: event - dict с httpMethod, queryStringParameters (format: xlsx/csv)
          context - объект с request_id
    Returns: HTTP response с файлом или ошибкой
    '''
    method: str = event.get('httpMethod', 'GET')
    
    if method == 'OPTIONS':
        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Methods': 'GET, OPTIONS',
                'Access-Control-Allow-Headers': 'Content-Type',
                'Access-Control-Max-Age': '86400'
            },
            'body': '',
            'isBase64Encoded': False
        }
    
    if method == 'GET':
        params = event.get('queryStringParameters', {})
        export_format = params.get('format', 'xlsx')
        
        db_url = os.environ.get('DATABASE_URL')
        conn = psycopg2.connect(db_url)
        cursor = conn.cursor()
        
        sql = """
            SELECT name, slug, subscribers, category, description, 
                   tme_link, language, posts_per_day, avg_views, 
                   contacts, last_checked, status
            FROM channels 
            ORDER BY subscribers DESC
        """
        
        cursor.execute(sql)
        rows = cursor.fetchall()
        
        if export_format == 'csv':
            import csv
            output = BytesIO()
            output.write('\ufeff'.encode('utf-8'))
            
            writer = csv.writer(output)
            writer.writerow([
                'Название', 'Slug', 'Подписчики', 'Категория', 'Описание',
                'Telegram-ссылка', 'Язык', 'Постов/день', 'Ср. просмотры',
                'Контакты', 'Проверен', 'Статус'
            ])
            
            for row in rows:
                writer.writerow([
                    row[0], row[1], row[2], row[3], row[4],
                    row[5], row[6], row[7], row[8], row[9],
                    row[10].strftime('%Y-%m-%d') if row[10] else '',
                    row[11]
                ])
            
            csv_data = output.getvalue()
            output.close()
            
            cursor.close()
            conn.close()
            
            return {
                'statusCode': 200,
                'headers': {
                    'Content-Type': 'text/csv; charset=utf-8',
                    'Content-Disposition': 'attachment; filename="tgstat_channels.csv"',
                    'Access-Control-Allow-Origin': '*'
                },
                'body': base64.b64encode(csv_data).decode('utf-8'),
                'isBase64Encoded': True
            }
        
        elif export_format == 'xlsx':
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Каналы TGStat"
                
                headers = [
                    'Название', 'Slug', 'Подписчики', 'Категория', 'Описание',
                    'Telegram-ссылка', 'Язык', 'Постов/день', 'Ср. просмотры',
                    'Контакты', 'Проверен', 'Статус'
                ]
                
                header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF', size=12)
                
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                for row_num, row_data in enumerate(rows, 2):
                    ws.cell(row=row_num, column=1, value=row_data[0])
                    ws.cell(row=row_num, column=2, value=row_data[1])
                    ws.cell(row=row_num, column=3, value=row_data[2])
                    ws.cell(row=row_num, column=4, value=row_data[3])
                    ws.cell(row=row_num, column=5, value=row_data[4])
                    ws.cell(row=row_num, column=6, value=row_data[5])
                    ws.cell(row=row_num, column=7, value=row_data[6])
                    ws.cell(row=row_num, column=8, value=float(row_data[7]) if row_data[7] else 0)
                    ws.cell(row=row_num, column=9, value=row_data[8])
                    ws.cell(row=row_num, column=10, value=row_data[9])
                    ws.cell(row=row_num, column=11, value=row_data[10].strftime('%Y-%m-%d') if row_data[10] else '')
                    ws.cell(row=row_num, column=12, value=row_data[11])
                
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column].width = adjusted_width
                
                output = BytesIO()
                wb.save(output)
                xlsx_data = output.getvalue()
                output.close()
                
                cursor.close()
                conn.close()
                
                return {
                    'statusCode': 200,
                    'headers': {
                        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        'Content-Disposition': 'attachment; filename="tgstat_channels.xlsx"',
                        'Access-Control-Allow-Origin': '*'
                    },
                    'body': base64.b64encode(xlsx_data).decode('utf-8'),
                    'isBase64Encoded': True
                }
            except ImportError:
                cursor.close()
                conn.close()
                
                return {
                    'statusCode': 500,
                    'headers': {
                        'Content-Type': 'application/json',
                        'Access-Control-Allow-Origin': '*'
                    },
                    'body': json.dumps({
                        'error': 'openpyxl не установлен. Используйте format=csv'
                    }),
                    'isBase64Encoded': False
                }
        
        cursor.close()
        conn.close()
        
        return {
            'statusCode': 400,
            'headers': {
                'Content-Type': 'application/json',
                'Access-Control-Allow-Origin': '*'
            },
            'body': json.dumps({
                'error': 'Неподдерживаемый формат. Используйте format=xlsx или format=csv'
            }),
            'isBase64Encoded': False
        }
    
    return {
        'statusCode': 405,
        'headers': {
            'Content-Type': 'application/json',
            'Access-Control-Allow-Origin': '*'
        },
        'body': json.dumps({'error': 'Метод не поддерживается'}),
        'isBase64Encoded': False
    }
